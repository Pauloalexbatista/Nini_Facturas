import os
import json
import time
import shutil
import pandas as pd
import google.generativeai as genai
from dotenv import load_dotenv
from openpyxl.styles import Font, Alignment
import traceback

# 1. SETUP E CONFIGURAÇÃO
load_dotenv()
api_key = os.getenv("GOOGLE_API_KEY")

if not api_key or api_key == "Cole_Aqui_A_Sua_Chave_Privada":
    print("ERRO: API Key não configurada no arquivo .env!")
    exit()

genai.configure(api_key=api_key)
model = genai.GenerativeModel('gemini-flash-latest')

# Caminhos
PASTA_ENTRADA = "1_Faturas_Entrada"
PASTA_PROCESSADOS = "2_Faturas_Processadas"
PASTA_SAIDA = "3_Ficheiro_Entradas"
TABELA_MAPEAMENTO = "tabela_mapeamento.xlsx"

def processar_faturas():
    # Garantir que as pastas existem
    for pasta in [PASTA_ENTRADA, PASTA_PROCESSADOS, PASTA_SAIDA]:
        if not os.path.exists(pasta):
            os.makedirs(pasta)
    
    print("\n" + "="*40)
    print("      CONFIGURAÇÃO DE ENTRADA")
    print("="*40)
    
    # Solicitar data de entrada (padrão hoje)
    hoje_ddmm = time.strftime("%d%m")
    hoje_ddmmyyyy = time.strftime("%d%m%Y")
    
    data_input = input(f"Digite a data de entrada (DDMM) [Vazio para hoje {hoje_ddmm}]: ").strip()
    
    if data_input:
        if len(data_input) == 4:
            entry_ddmm = data_input
            entry_ddmmyyyy = f"{data_input}{time.strftime('%Y')}"
        else:
            print("Formato inválido! Usando data de hoje.")
            entry_ddmm = hoje_ddmm
            entry_ddmmyyyy = hoje_ddmmyyyy
    else:
        entry_ddmm = hoje_ddmm
        entry_ddmmyyyy = hoje_ddmmyyyy

    print(f"Data selecionada: {entry_ddmm}/{time.strftime('%Y')}")
    # 2. CARREGAR MAPEAMENTO
    if not os.path.exists(TABELA_MAPEAMENTO):
        print(f"ERRO: Arquivo {TABELA_MAPEAMENTO} não encontrado!")
        return
    
    try:
        df_artigos = pd.read_excel(TABELA_MAPEAMENTO, sheet_name='Artigos')
        df_fornecedores = pd.read_excel(TABELA_MAPEAMENTO, sheet_name='Fornecedores')
    except Exception as e:
        print(f"Erro ao ler abas do Excel: {e}")
        return
    
    def limpar_id(val):
        s = str(val).strip()
        if s.endswith('.0'): return s[:-2]
        return s

    def normalize_name(name):
        import re
        if not name: return ""
        # Converte para upper e remove sufixos comuns
        n = str(name).upper()
        n = re.sub(r'\b(LDA|SA|SL|UNIPESSOAL|LIMITADA)\b', '', n)
        # Remove caracteres especiais
        return re.sub(r'[^A-Z0-9]', '', n)

    df_artigos['Cod_Artigo_Fornecedor'] = df_artigos['Cod_Artigo_Fornecedor'].astype(str).str.strip()
    df_artigos['Fornecedor_ID'] = df_artigos['Fornecedor_ID'].apply(limpar_id)
    df_fornecedores['Fornecedor_ID'] = df_fornecedores['Fornecedor_ID'].apply(limpar_id)
    
    # Normaliza todos os nomes possíveis (até 3)
    df_fornecedores['Norm1'] = ""
    df_fornecedores['Norm2'] = ""
    df_fornecedores['Norm3'] = ""
    
    if 'Nome_Fornecedor1' in df_fornecedores.columns:
        df_fornecedores['Norm1'] = df_fornecedores['Nome_Fornecedor1'].apply(normalize_name)
    if 'Nome_Fornecedor2' in df_fornecedores.columns:
        df_fornecedores['Norm2'] = df_fornecedores['Nome_Fornecedor2'].apply(normalize_name)
    if 'Nome_Fornecedor3' in df_fornecedores.columns:
        df_fornecedores['Norm3'] = df_fornecedores['Nome_Fornecedor3'].apply(normalize_name)

    resultados = []
    sucesso_count = 0
    erro_count = 0
    pdfs = [f for f in os.listdir(PASTA_ENTRADA) if f.lower().endswith('.pdf')]
    
    if not pdfs:
        print(f"\nNenhuma fatura encontrada na pasta '{PASTA_ENTRADA}'.")
        input("\nPressione Enter para fechar...")
        return

    for pdf_name in pdfs:
        caminho_pdf = os.path.join(PASTA_ENTRADA, pdf_name)
        print(f"\n--- Lendo: {pdf_name} ---")
        
        try:
            pdf_file = genai.upload_file(caminho_pdf)
            prompt = """
            Analise esta fatura de peixe e extraia os itens em formato JSON.
            Campos por item: Cod_Artigo_Fornecedor, Nome_Artigo, Peso_KG, Quantidade, Lote, Preco_Unitario, Info_Tecnica.
            
            No campo 'Info_Tecnica', agrupe de forma resumida: 
            - Origem (País)
            - Capturado ou Aquacultura
            - Método de Pesca
            - Estado (Fresco ou Congelado)
            
            Também identifique:
            1. O NOME ou ID do FORNECEDOR da fatura.
            2. A DATA da fatura (no formato DD/MM/AAAA).
            
            Responda obrigatoriamente neste formato JSON:
            {
              "fornecedor_detectado": "Nome do Fornecedor",
              "data_fatura": "DD/MM/AAAA",
              "itens": [
                {"Cod_Artigo_Fornecedor": "...", "Nome_Artigo": "...", "Peso_KG": 0, "Preco_Unitario": 0, "Quantidade": 0, "Lote": "...", "Info_Tecnica": "..."},
                ...
              ]
            }
            Se não encontrar um campo, coloque nulo (null). Responda APENAS com o JSON.
            """
            
            response = model.generate_content([prompt, pdf_file])
            cleaned_response = response.text.strip().replace("```json", "").replace("```", "")
            data_extracted = json.loads(cleaned_response)
            
            fornecedor_txt = data_extracted.get('fornecedor_detectado', '')
            itens = data_extracted.get('itens', [])

            for item in itens:
                cod_art_forn = str(item.get('Cod_Artigo_Fornecedor', '')).strip()
                match_art = df_artigos[df_artigos['Cod_Artigo_Fornecedor'] == cod_art_forn]
                
                # Setup padrão
                forn_id = ""
                abrev = "??"
                nome_forn = fornecedor_txt
                classif = ""
                tipo_ent = ""
                doc_sist = "99_OUTROS"
                pesq_forn_name = normalize_name(fornecedor_txt)

                if not match_art.empty:
                    row_art = match_art.iloc[0]
                    item['Meu_Cod_Interno'] = row_art.get('Meu_Cod_Interno', 'NÃO MAPEADO')
                    item['Metodo_Producao'] = row_art.get('Metodo_Producao', '')
                    item['Zona_Mar'] = row_art.get('Zona_Mar', '')
                    item['Massa_Agua_Origem'] = row_art.get('Massa_Agua_Origem', '')
                    item['Origem_Aquacultura'] = row_art.get('Origem_Aquacultura', '')
                    item['Arte_Pesca'] = row_art.get('Arte_Pesca', '')
                    item['Zona_Pesca'] = row_art.get('Zona_Pesca', '')
                    item['Metodo_Pesca'] = row_art.get('Metodo_Pesca', '')
                    item['Origem'] = row_art.get('Origem', '')
                    item['Estado'] = row_art.get('Estado', '')
                    f_id = str(row_art.get('Fornecedor_ID', '')).strip()
                    match_forn = df_fornecedores[df_fornecedores['Fornecedor_ID'] == f_id]
                else:
                    item['Meu_Cod_Interno'] = "NÃO MAPEADO"
                    # Inicializa campos de mapeamento (como no match_art) para evitar erros
                    map_fields = ['Metodo_Producao', 'Zona_Mar', 'Massa_Agua_Origem', 
                                  'Origem_Aquacultura', 'Arte_Pesca', 'Zona_Pesca', 
                                  'Metodo_Pesca', 'Origem', 'Estado']
                    for f in map_fields: item[f] = ""
                    
                    # Procura o fornecedor pelo nome detectado
                    match_forn = df_fornecedores[
                        (df_fornecedores['Norm1'] == pesq_forn_name) | 
                        (df_fornecedores['Norm2'] == pesq_forn_name) | 
                        (df_fornecedores['Norm3'] == pesq_forn_name)
                    ]

                if not match_forn.empty:
                    row_forn = match_forn.iloc[0]
                    forn_id = row_forn.get('Fornecedor_ID', '')
                    nome_forn = row_forn.get('Nome_Fornecedor1', row_forn.get('Nome_Fornecedor', nome_forn))
                    abrev = str(row_forn.get('Abreviatura', '??'))
                    classif = row_forn.get('Classificacao', '')
                    tipo_ent = row_forn.get('Tipo_Entrada', '')
                    doc_sist = row_forn.get('Doc_Entrada_sistema', '99_OUTROS')
                
                # Consolidação de Características para o layout final de 4 colunas (Layout compactado)
                item['Zona_Pesca_Output'] = item.get('Zona_Mar', '') or item.get('Zona_Pesca', '')
                item['Metodo_Pesca_Output'] = item.get('Arte_Pesca', '') or item.get('Metodo_Producao', '') or item.get('Metodo_Pesca', '')
                item['Origem_Output'] = item.get('Origem_Aquacultura', '') or item.get('Massa_Agua_Origem', '') or item.get('Origem', '')
                item['Estado_Output'] = item.get('Estado', '')

                # REGRA DO LOTE
                lote_forn = item.get('Lote')
                if lote_forn and lote_forn != "null" and lote_forn is not None:
                    item['Lote'] = f"{abrev}.{entry_ddmm}.{lote_forn}"
                else:
                    item['Lote'] = f"{abrev}.{entry_ddmmyyyy}"

                item['Fornecedor_ID'] = forn_id
                item['Nome_Fornecedor'] = nome_forn
                item['Abreviatura'] = abrev
                item['Classificacao'] = classif
                item['Tipo_Entrada'] = tipo_ent
                item['Doc_Entrada_sistema'] = doc_sist
                item['Preco'] = item.get('Preco_Unitario', 0)
                item['Arquivo_Original'] = pdf_name
                resultados.append(item)

            shutil.move(caminho_pdf, os.path.join(PASTA_PROCESSADOS, pdf_name))
            print(f"Sucesso! {pdf_name} processado.")
            sucesso_count += 1

        except Exception as e:
            print(f"Erro ao processar {pdf_name}: {str(e)[:100]}")
            erro_count += 1

    if resultados:
        df_final = pd.DataFrame(resultados)
        
        # Ordenar antes de exportar
        if 'Doc_Entrada_sistema' in df_final.columns and 'Nome_Fornecedor' in df_final.columns:
            df_final = df_final.sort_values(by=['Doc_Entrada_sistema', 'Nome_Fornecedor'])

        arquivo_nome = f"importar_para_sistema_{time.strftime('%Y%m%d_%H%M%S')}.xlsx"
        arquivo_saida = os.path.join(PASTA_SAIDA, arquivo_nome)

        # EXPORTAÇÃO CUSTOMIZADA (LAYOUT DO USER)
        try:
            with pd.ExcelWriter(arquivo_saida, engine='openpyxl') as writer:
                # Se o pandas criar uma planilha 'Sheet1', vamos removê-la ou renomear
                # Mas é mais limpo criar o workbook e a sheet do zero via openpyxl se necessário.
                # Aqui vamos usar o writer.book que é um Workbook do openpyxl.
                workbook = writer.book
                if 'Sheet1' in workbook.sheetnames:
                    std_sheet = workbook['Sheet1']
                    std_sheet.title = "Importação"
                else:
                    workbook.create_sheet("Importação")
                
                worksheet = workbook["Importação"]
                
                # Ajustar Largura das Colunas
                worksheet.column_dimensions['A'].width = 10  # CODIGO
                worksheet.column_dimensions['B'].width = 10  # Peso
                worksheet.column_dimensions['C'].width = 20  # Lote
                worksheet.column_dimensions['D'].width = 50  # Info_Tecnica (Larga mas com wrap)
                worksheet.column_dimensions['E'].width = 10  # Preco
                
                curr_row = 1
                
                # Garantir que todos os campos necessários existem
                required_cols = ['Doc_Entrada_sistema', 'Fornecedor_ID', 'Abreviatura', 'Nome_Fornecedor', 'Classificacao']
                for col in required_cols:
                    if col not in df_final.columns:
                        df_final[col] = ""

                # Agrupar por documento e fornecedor
                grupos = df_final.groupby(['Doc_Entrada_sistema', 'Fornecedor_ID', 'Abreviatura', 'Nome_Fornecedor', 'Classificacao'], sort=False)
                
                for (doc_sist, forn_id, abrev, nome_forn, classif), group_data in grupos:
                    # 1. Header do Documento e Fornecedor (Compacto)
                    worksheet.cell(row=curr_row, column=2, value='Doc_Entrada_sistema')
                    worksheet.cell(row=curr_row, column=3, value='Fornecedor_ID')
                    worksheet.cell(row=curr_row, column=4, value='Nome_Fornecedor')
                    worksheet.cell(row=curr_row, column=5, value='Abreviatura')
                    worksheet.cell(row=curr_row, column=6, value='Classificacao')
                    
                    worksheet.cell(row=curr_row+1, column=2, value=doc_sist)
                    worksheet.cell(row=curr_row+1, column=3, value=forn_id)
                    worksheet.cell(row=curr_row+1, column=4, value=nome_forn)
                    worksheet.cell(row=curr_row+1, column=5, value=abrev)
                    worksheet.cell(row=curr_row+1, column=6, value=classif)
                    
                    # 2. Header dos Itens - Reduzido e Reordenado (Layout 9 colunas final)
                    item_headers = ['CODIGO', 'Peso_KG', 'Lote', 'Info_Tecnica', 'Preco',
                                    'Zona_Pesca', 'Metodo_Pesca', 'Origem', 'Estado']
                    for col_idx, header in enumerate(item_headers, start=1):
                        cell = worksheet.cell(row=curr_row+2, column=col_idx, value=header)
                        # Cabeçalho em negrito
                        cell.font = Font(bold=True)
                    
                    # 4. Dados dos Itens
                    for i, (_, row) in enumerate(group_data.iterrows()):
                        # CODIGO
                        worksheet.cell(row=curr_row+3+i, column=1, value=row.get('Meu_Cod_Interno'))
                        # Peso_KG
                        c_peso = worksheet.cell(row=curr_row+3+i, column=2, value=row.get('Peso_KG'))
                        c_peso.number_format = '#,##0.00'
                        # Lote
                        worksheet.cell(row=curr_row+3+i, column=3, value=row.get('Lote'))
                        # Info_Tecnica
                        c_info = worksheet.cell(row=curr_row+3+i, column=4, value=row.get('Info_Tecnica'))
                        c_info.alignment = Alignment(wrap_text=True)
                        # Preco
                        c_preco = worksheet.cell(row=curr_row+3+i, column=5, value=row.get('Preco'))
                        c_preco.number_format = '#,##0.00'
                        # Características Consolidadas
                        worksheet.cell(row=curr_row+3+i, column=6, value=row.get('Zona_Pesca_Output'))
                        worksheet.cell(row=curr_row+3+i, column=7, value=row.get('Metodo_Pesca_Output'))
                        worksheet.cell(row=curr_row+3+i, column=8, value=row.get('Origem_Output'))
                        worksheet.cell(row=curr_row+3+i, column=9, value=row.get('Estado_Output'))
                    
                    # Espaço entre grupos (menos espaço para economizar papel)
                    curr_row += 3 + len(group_data) + 1

        except Exception as e:
            print(f"Erro ao gerar Excel customizado: {e}")
            # Fallback para o modo simples se der erro cabuloso
            df_final.to_excel(arquivo_saida, index=False)
            print("Gerado Excel em formato simples (backup).")
        
        print("\n" + "="*40)
        print("           RELATÓRIO FINAL")
        print("="*40)
        print(f"Faturas processadas com sucesso: {sucesso_count}")
        print(f"Faturas com erro:               {erro_count}")
        print(f"Arquivo gerado: {arquivo_saida}")
        print("="*40)
    else:
        print("\nNenhum dado extraído das faturas.")
    
    input("\nPressione Enter para fechar a janela...")

if __name__ == "__main__":
    try:
        processar_faturas()
    except Exception as e:
        print("\n" + "!"*40)
        print("      ERRO CRÍTICO NO PROGRAMA")
        print("!"*40)
        traceback.print_exc()
        input("\nPressione Enter para fechar a janela...")
